import urllib
import os
import json
import datetime

import requests
from openpyxl import load_workbook

def get_file_name(url):
    path=urllib.parse.urlparse(url).path
    return os.path.split(path)[-1]

def download_osaka_model_data_file():
    #大阪府の大阪モデルモニタリング指標のページにある　Excelファイルのリンク
    url="http://www.pref.osaka.lg.jp/attach/23711/00362734/sihyou.xlsx"
    excel_file_name=get_file_name(url)

    response=requests.get(url)
    if response.status_code!=requests.codes.ok:
        raise Exception("status_code!=200")

    print(excel_file_name)
    with open(excel_file_name,"wb") as f:
        f.write(response.content)
    return excel_file_name


def get_bed_data_for_mild_or_moderate_patients(ws):
    data=[]
    found_first_value=None
    for col in range(2,ws.max_column+1):

        dt=ws.cell(37,col).value
        num_beds=ws.cell(38,col).value

        #最初の部分で未記入のものは飛ばす
        if num_beds is None and not found_first_value:
            continue
            
        #最新の未記入のデータの部分で終了
        if num_beds is None and found_first_value:
            break 

        num_patients=ws.cell(39,col).value
        ratio=ws.cell(40,col).value
        print([dt,num_beds,num_patients,ratio])

        data.append({
            "date":dt.date().isoformat(),
            "hospital_capacity":num_beds,
            "num_patients":num_patients,
            "ratio":ratio
        })

    return data


def get_bed_data_for_severe_patients(ws):
    data=[]
    found_first_value=None
    for col in range(2,ws.max_column+1):

        dt=ws.cell(31,col).value
        num_beds=ws.cell(32,col).value

        #最初の部分で未記入のものは飛ばす
        if num_beds is None and not found_first_value:
            continue
            
        #最新の未記入のデータの部分で終了
        if num_beds is None and found_first_value:
            break 

        num_patients=ws.cell(33,col).value
        ratio=ws.cell(34,col).value
        print([dt,num_beds,num_patients,ratio])

        data.append({
            "date":dt.date().isoformat(),
            "hospital_capacity":num_beds,
            "num_patients":num_patients,
            "ratio":ratio
        })

    return data


def get_accommodation_facility_data(ws):
    data=[]
    found_first_value=None
    for col in range(2,ws.max_column+1):
        dt=ws.cell(43,col).value
        num_rooms=ws.cell(44,col).value

        #最初の部分で未記入のものは飛ばす
        if num_rooms is None and not found_first_value:
            continue
            
        #最新の未記入のデータの部分で終了
        if num_rooms is None and found_first_value:
            break 

        num_patients=ws.cell(45,col).value
        ratio=ws.cell(46,col).value
        print([dt,num_rooms,num_patients,ratio])

        data.append({
            "date":dt.date().isoformat(),
            "num_rooms":num_rooms,
            "num_patients":num_patients,
            "ratio":ratio
        })

    return data


def main():
    file_name=download_osaka_model_data_file()


    wb=load_workbook(file_name,data_only=True)
    ws=wb["データ一覧"]

    bed_data_for_severe=get_bed_data_for_severe_patients(ws)
    bed_data_for_mild_moderate=get_bed_data_for_mild_or_moderate_patients(ws)
    data_accommodation_facility=get_accommodation_facility_data(ws)

    data={
        "data":{
            "severe":bed_data_for_severe,
            "mild_moderate":bed_data_for_mild_moderate,
            "accommodation_facility":data_accommodation_facility
        }
    }

    with open("osaka_bed_data.json","w") as f:
        json.dump(data,f,indent=4)

if __name__ == "__main__":
    main()
